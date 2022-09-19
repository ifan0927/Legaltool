from openpyxl import Workbook ,styles, worksheet , load_workbook
from openpyxl.styles.borders import Border, Side
from openpyxl.utils import get_column_letter
import sys,os

wb = Workbook()
ws0 = wb.create_sheet("繼承系統表",0)

## 全邊框
thin_border = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thin'), 
                     bottom=Side(style='thin'))

## 橫border(靠下)
bottom_border = Border(bottom=Side(style='thin'))
right_border = Border(right=Side(style='thin'))

filename = '繼承系統表.xlsx'

## excel內style設定
font_title = styles.Font(u'標楷體',size=16,bold=True)
font_context = styles.Font(u'標楷體',size=11)


def nums(first_number, last_number, step=1):
    return range(first_number, last_number+1, step)

class person:
    
    def __init__(self, check, bday , dday, wife, parent, pid,call):
        self.check = check
        self.bday = bday
        self.dday = dday
        self.wife = wife
        self.parent = parent
        self.pid = pid
        self.call = call
        self.soncount = 0
        self.drawed = 0
        self.s_drawed = 0
        self.w_drawed = 0

    def gen_set(self, int):
        self.gen = int

    def draw_self(self, s_row,s_col,name):
        ws0.merge_cells(start_row = s_row + 1,start_column = s_col, end_row = s_row + 2, end_column= s_col)
        ws0.cell(row = s_row, column = s_col).fill = styles.PatternFill("solid", fgColor="000000")
        ws0.cell(row = s_row, column = s_col + 1).value = name
        ws0.cell(row = s_row + 1, column = s_col).value = self.call
        ws0.cell(row = s_row + 1, column = s_col + 1).value = self.bday
        ws0.cell(row = s_row + 2, column = s_col + 1).value = self.dday
        if (self.check == 1):
            ws0.cell(row = s_row + 3, column= s_col ).value = "有拋棄繼承"
        else: 
            ws0.cell(row = s_row + 3, column= s_col ).value = "無拋棄繼承"
        ws0.cell(row = s_row + 3, column= s_col + 1).value = self.pid
        for c in nums(s_col, s_col + 1):
            for r in nums(s_row , s_row + 3):
                ws0.cell(row= r , column=c).font  = font_context
                ws0.cell(row= r , column=c).alignment = styles.Alignment(horizontal="center",vertical='center')
                ws0.cell(row= r , column=c).border =  thin_border
        self.drawed = 1
        
    
    # def draw_son(self, s_col,s_row, son, dict):

    #     ws0.cell(row = s_row + 1 , column = s_col+2).border =  bottom_border
    #     ws0.cell(row = s_row + 1 , column = s_col+3).border =  bottom_border
    #     dict[son[0]].draw_self(s_col +4, s_row , son[0])
    #     if len(son) > 1:
    #         ##多個兒子的狀況
    #         print('TODO')
    #     else:
    #         #row+1 col+2/col+3 畫兩個底線位置,
            
        
    #     self.s_drawed = 1
            
        

def make_excel(d, l, app, gen): ## (存放全部object的dict,全體資料名稱的list, flask app, 全體資料gen的list)
    max_gen = max(gen)
    
    #excel檔內標題 字型/標題
    ws0.merge_cells('A1:C4') 
    ws0['A1'] = l[0] + "的繼承系統表"
    ws0.cell(row=1,column=1).alignment = styles.Alignment(horizontal="center",vertical='center') ## 置中對齊
    ws0.cell(row=1,column=1).font = font_title ##標題字形

    ws0['B5'] =  "表格範例"
    ws0.cell(row=5,column=2).alignment = styles.Alignment(horizontal="center",vertical='center')
    ws0.cell(row=5 , column=2).font  = font_context


    ##excel檔內範例 內容/字型/標題
    ws0.merge_cells("B7:B8")
    ws0['B6'] = '是否為繼承人,若為繼承人為此格為黑底'
    ws0['C6'] = '姓名'
    ws0['B7'] = '稱謂'
    ws0['C7'] = '出生日'
    ws0['C8'] = '死亡日'
    ws0['B9'] = '有無拋棄繼承'
    ws0['C9'] = '身分證字號'
    for r in nums(6,9):
        for c in nums(2,3):
            ws0.cell(row= r , column=c).font  = font_context
            ws0.cell(row= r , column=c).alignment = styles.Alignment(horizontal="center",vertical='center')
            ws0.cell(row= r , column=c).border =  thin_border
    ws0.column_dimensions['A'].width = 1
    ws0.column_dimensions['B'].width = 42
    ws0.column_dimensions['C'].width = 14 
    ws0.column_dimensions['D'].width = 1 
    
    for i in range(max_gen):
        A = get_column_letter(5 + 4 * i)
        B = get_column_letter(6 + 4 * i)
        C = get_column_letter(7 + 4 * i)
        D = get_column_letter(8 + 4 * i)
        ws0.column_dimensions[str(A)].width = 13
        ws0.column_dimensions[str(B)].width = 14
        ws0.column_dimensions[str(C)].width = 4
        ws0.column_dimensions[str(D)].width = 4

    ##先畫出最後一代
    draw_lastgen(d, l, gen, max_gen)
    
    wb.save(app.root_path +'/xlsx/'+ filename)

## (row, col) , 第一代起始:(1,5) (1,9) ......(1,5 + 4*(gen-1)) 

def draw_lastgen(d, l, gen, max_gen):
    parentgroup = dict() ##字典以各parent當key,將資料分組
    thisgen = list() ##紀錄這一代有誰
    s_row = 1
    s_col = 5 + 4 * (max_gen - 1)

    for x in range(len(l)):
        if (gen[x] == max_gen):
            thisgen.append(x) 
            if parentgroup.__contains__(d[l[x]].parent):
                parentgroup[d[l[x]].parent].append(x)
            else:
                parentgroup[d[l[x]].parent] = list()
                parentgroup[d[l[x]].parent].append(x)
    
    count = 0
    c = 0
    for x in parentgroup:     
        for i in range(len(parentgroup[x])):
            index = parentgroup[x][i]
            if (count == 0):           
                d[l[index]].draw_self(s_row,s_col,l[index])
                count = count + 1
            else:
                d[l[index]].draw_self(s_row + (c+ 1) * 5, s_col, l[index])
                count = count + 1
                c = c + 1
            
        
